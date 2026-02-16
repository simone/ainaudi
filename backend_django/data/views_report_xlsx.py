"""
View per generare report XLSX della mappatura RDL per comune.
"""
import io
from collections import defaultdict

from django.http import HttpResponse
from rest_framework import permissions
from rest_framework.response import Response
from rest_framework.views import APIView

from core.permissions import CanManageMappatura
from .models import SectionAssignment
from campaign.models import RdlRegistration
from territory.models import SezioneElettorale
from delegations.models import DesignazioneRDL
from data.views import get_consultazione_attiva, resolve_comune_id


class MappaturaReportXlsxView(APIView):
    """
    Download XLSX report of sezioni in a comune with RDL assignments.

    GET /api/mappatura/report-xlsx/?comune_id=X
        &sezione_ids=1,2,3        (optional: only these sezioni)
        &includi_confermati=true   (optional: include confirmed designations, default true)

    Returns XLSX with columns:
    STATO | COMUNE | MUNICIPIO | SEZIONE | INDIRIZZO | DENOMINAZIONE |
    EFFETTIVO COGNOME | EFFETTIVO NOME | EFFETTIVO DATA NASCITA | EFFETTIVO DOMICILIO |
    SUPPLENTE COGNOME | SUPPLENTE NOME | SUPPLENTE DATA NASCITA | SUPPLENTE DOMICILIO
    """
    permission_classes = [permissions.IsAuthenticated, CanManageMappatura]

    def post(self, request):
        from delegations.permissions import get_sezioni_filter_for_user, has_referenti_permission
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        consultazione = get_consultazione_attiva()
        if not consultazione:
            return Response({'error': 'Nessuna consultazione attiva'}, status=400)

        if not has_referenti_permission(request.user, consultazione.id):
            return Response({'error': 'Non autorizzato'}, status=403)

        comune_param = request.data.get('comune_id')
        if not comune_param:
            return Response({'error': 'comune_id è obbligatorio'}, status=400)

        comune_id = resolve_comune_id(comune_param)
        if not comune_id:
            return Response({'error': 'Comune non trovato'}, status=404)

        # Optional: filter to specific sezione IDs (sent in body, no URL limit)
        specific_ids = request.data.get('sezione_ids')
        if specific_ids and not isinstance(specific_ids, list):
            specific_ids = None

        includi_confermati = request.data.get('includi_confermati', True)

        # Get user's allowed sezioni filter
        sezioni_filter = get_sezioni_filter_for_user(request.user, consultazione.id)
        if sezioni_filter is None:
            return Response({'error': 'Non autorizzato per questo territorio'}, status=403)

        # Query ALL sezioni in the comune
        from django.db.models import Q
        filters = Q(is_attiva=True) & sezioni_filter & Q(comune_id=comune_id)
        if specific_ids:
            filters &= Q(id__in=specific_ids)

        sezioni = SezioneElettorale.objects.filter(filters).select_related(
            'comune', 'municipio'
        ).order_by('municipio__numero', 'numero')

        sezioni_ids = [s.id for s in sezioni]

        # Get mappatura assignments (SectionAssignment)
        assignments = SectionAssignment.objects.filter(
            sezione_id__in=sezioni_ids,
            consultazione=consultazione,
        ).select_related('rdl_registration')

        assignment_map = defaultdict(dict)
        for a in assignments:
            assignment_map[a.sezione_id][a.role] = a.rdl_registration

        # Get confirmed designations
        confermati_map = {}
        if includi_confermati:
            designazioni = DesignazioneRDL.objects.filter(
                sezione_id__in=sezioni_ids,
                stato='CONFERMATA',
                is_attiva=True,
            )
            for d in designazioni:
                confermati_map[d.sezione_id] = d

        # Get comune name for filename
        comune_nome = sezioni[0].comune.nome if sezioni else 'comune'

        # Build rows: merge mappatura + confermati
        rows = []
        sezioni_seen = set()

        def domicilio_reg(reg):
            """Domicilio from RdlRegistration (priorità domicilio > residenza)."""
            if not reg:
                return ''
            if reg.comune_domicilio and reg.indirizzo_domicilio:
                return f"{reg.indirizzo_domicilio}, {reg.comune_domicilio}"
            if reg.indirizzo_residenza and reg.comune_residenza:
                return f"{reg.indirizzo_residenza}, {reg.comune_residenza}"
            return ''

        for sez in sezioni:
            roles = assignment_map.get(sez.id, {})
            desig = confermati_map.get(sez.id)

            # Determine stato and data source
            if desig:
                stato = 'CONFERMATO'
                # Use snapshot data from designazione
                row = [
                    stato,
                    sez.comune.nome,
                    f"Municipio {sez.municipio.numero}" if sez.municipio else '',
                    sez.numero,
                    sez.indirizzo or '',
                    sez.denominazione or '',
                    desig.effettivo_cognome or '',
                    desig.effettivo_nome or '',
                    desig.effettivo_data_nascita.strftime('%d/%m/%Y') if desig.effettivo_data_nascita else '',
                    desig.effettivo_domicilio or '',
                    desig.supplente_cognome or '',
                    desig.supplente_nome or '',
                    desig.supplente_data_nascita.strftime('%d/%m/%Y') if desig.supplente_data_nascita else '',
                    desig.supplente_domicilio or '',
                ]
                rows.append(row)
                sezioni_seen.add(sez.id)
            elif roles:
                stato = 'MAPPATO'
                eff = roles.get('RDL')
                sup = roles.get('SUPPLENTE')
                row = [
                    stato,
                    sez.comune.nome,
                    f"Municipio {sez.municipio.numero}" if sez.municipio else '',
                    sez.numero,
                    sez.indirizzo or '',
                    sez.denominazione or '',
                    eff.cognome if eff else '',
                    eff.nome if eff else '',
                    eff.data_nascita.strftime('%d/%m/%Y') if eff and eff.data_nascita else '',
                    domicilio_reg(eff),
                    sup.cognome if sup else '',
                    sup.nome if sup else '',
                    sup.data_nascita.strftime('%d/%m/%Y') if sup and sup.data_nascita else '',
                    domicilio_reg(sup),
                ]
                rows.append(row)
                sezioni_seen.add(sez.id)

        # Build XLSX
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Mappatura RDL'

        # Styles
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
        confermato_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        mappato_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

        headers = [
            'STATO', 'COMUNE', 'MUNICIPIO', 'SEZIONE', 'INDIRIZZO', 'DENOMINAZIONE',
            'EFFETTIVO COGNOME', 'EFFETTIVO NOME', 'EFFETTIVO DATA NASCITA', 'EFFETTIVO DOMICILIO',
            'SUPPLENTE COGNOME', 'SUPPLENTE NOME', 'SUPPLENTE DATA NASCITA', 'SUPPLENTE DOMICILIO',
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows
        for row_idx, row_data in enumerate(rows, 2):
            stato = row_data[0]
            fill = confermato_fill if stato == 'CONFERMATO' else mappato_fill
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                cell.fill = fill

        # Auto-width columns
        for col in range(1, len(headers) + 1):
            max_length = len(str(ws.cell(row=1, column=col).value))
            for row in range(2, len(rows) + 2):
                val = ws.cell(row=row, column=col).value
                if val:
                    max_length = max(max_length, len(str(val)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(max_length + 3, 40)

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Write to response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"mappatura_rdl_{comune_nome.replace(' ', '_').lower()}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
