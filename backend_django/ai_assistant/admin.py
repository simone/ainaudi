"""
Django Admin configuration for AI Assistant models.
"""
from django.contrib import admin
from django.utils.translation import gettext_lazy as _
from django.http import HttpResponse
from .models import KnowledgeSource, ChatSession, ChatMessage
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


@admin.register(KnowledgeSource)
class KnowledgeSourceAdmin(admin.ModelAdmin):
    list_display = ['title', 'source_type', 'is_active', 'updated_at']
    list_filter = ['source_type', 'is_active']
    search_fields = ['title', 'content']
    readonly_fields = ['created_at', 'updated_at']


class ChatMessageInline(admin.TabularInline):
    model = ChatMessage
    extra = 0
    readonly_fields = ['role', 'content', 'created_at']


@admin.register(ChatSession)
class ChatSessionAdmin(admin.ModelAdmin):
    list_display = ['id', 'user_email', 'context', 'sezione', 'created_at', 'updated_at']
    list_filter = ['context', 'created_at']
    search_fields = ['user_email']
    raw_id_fields = ['sezione']
    inlines = [ChatMessageInline]
    readonly_fields = ['created_at', 'updated_at']
    actions = ['export_to_xlsx']

    def export_to_xlsx(self, request, queryset):
        """
        Export selected ChatSessions to XLSX with full message threads.
        Format: One row per message with session metadata.
        """
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Chat Sessions"

        # Define headers
        headers = [
            'Session ID',
            'User Email',
            'Context',
            'Sezione',
            'Session Title',
            'Session Created',
            'Session Updated',
            'Metadata',
            'Message #',
            'Message Role',
            'Message Content',
            'Message Created',
        ]

        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Populate data
        row_num = 2
        for session in queryset.prefetch_related('messages').order_by('-created_at'):
            messages = session.messages.order_by('created_at')

            # Get sezione info
            sezione_str = ''
            if session.sezione:
                sezione_str = f"{session.sezione.numero} - {session.sezione.comune.nome}"
                if session.sezione.municipio:
                    sezione_str += f" ({session.sezione.municipio.nome})"

            # Convert metadata to string
            metadata_str = str(session.metadata) if session.metadata else ''

            if messages.exists():
                # One row per message
                for msg_num, message in enumerate(messages, 1):
                    ws.cell(row=row_num, column=1, value=str(session.id))
                    ws.cell(row=row_num, column=2, value=session.user_email)
                    ws.cell(row=row_num, column=3, value=session.context or '')
                    ws.cell(row=row_num, column=4, value=sezione_str)
                    ws.cell(row=row_num, column=5, value=session.title or '')
                    ws.cell(row=row_num, column=6, value=session.created_at.strftime('%Y-%m-%d %H:%M:%S'))
                    ws.cell(row=row_num, column=7, value=session.updated_at.strftime('%Y-%m-%d %H:%M:%S'))
                    ws.cell(row=row_num, column=8, value=metadata_str)
                    ws.cell(row=row_num, column=9, value=msg_num)
                    ws.cell(row=row_num, column=10, value=message.role)
                    ws.cell(row=row_num, column=11, value=message.content)
                    ws.cell(row=row_num, column=12, value=message.created_at.strftime('%Y-%m-%d %H:%M:%S'))
                    row_num += 1
            else:
                # Session with no messages - single row
                ws.cell(row=row_num, column=1, value=str(session.id))
                ws.cell(row=row_num, column=2, value=session.user_email)
                ws.cell(row=row_num, column=3, value=session.context or '')
                ws.cell(row=row_num, column=4, value=sezione_str)
                ws.cell(row=row_num, column=5, value=session.title or '')
                ws.cell(row=row_num, column=6, value=session.created_at.strftime('%Y-%m-%d %H:%M:%S'))
                ws.cell(row=row_num, column=7, value=session.updated_at.strftime('%Y-%m-%d %H:%M:%S'))
                ws.cell(row=row_num, column=8, value=metadata_str)
                ws.cell(row=row_num, column=9, value=0)
                ws.cell(row=row_num, column=10, value='')
                ws.cell(row=row_num, column=11, value='(No messages)')
                ws.cell(row=row_num, column=12, value='')
                row_num += 1

        # Auto-size columns
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = len(str(headers[col_num - 1]))

            # Don't auto-size content column (11) as it can be very long
            if col_num != 11:
                for cell in ws[column_letter]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
            else:
                adjusted_width = 80  # Fixed width for message content

            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=chat_sessions_export.xlsx'

        # Save workbook to response
        wb.save(response)

        return response

    export_to_xlsx.short_description = "Esporta in XLSX (con messaggi completi)"


@admin.register(ChatMessage)
class ChatMessageAdmin(admin.ModelAdmin):
    list_display = ['session', 'role', 'content_preview', 'created_at']
    list_filter = ['role', 'created_at']
    search_fields = ['content']
    raw_id_fields = ['session']
    readonly_fields = ['created_at']

    def content_preview(self, obj):
        return obj.content[:100] + '...' if len(obj.content) > 100 else obj.content
    content_preview.short_description = _('Contenuto')
